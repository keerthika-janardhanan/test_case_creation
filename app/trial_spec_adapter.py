import json
import logging
import os
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Optional, Tuple, List

logger = logging.getLogger(__name__)

# Note: do not hardcode any example TestCaseID here. We will infer the case id
# from the spec file when available, or use the first credentials row as a
# neutral fallback.


@dataclass
class TrialCredentials:
    base_url: str
    username: str
    password: str


def _extract_titles_from_source(source: str) -> List[str]:
    """Return unique test titles found in a Playwright spec source."""
    titles: List[str] = []
    pattern = re.compile(r"\b(?:run|test)\(\s*['\"]([^'\"]+)['\"]")
    for m in pattern.finditer(source or ""):
        t = m.group(1).strip()
        if t and t not in titles:
            titles.append(t)
    return titles


def _is_id_like(value: str) -> bool:
    # Heuristic for ID-like tokens: contains underscore or all-caps or starts with 'TC'
    if not value:
        return False
    if "_" in value:
        return True
    if value.upper() == value and any(c.isalpha() for c in value):
        return True
    if value.startswith("TC"):
        return True
    return False


def load_trial_credentials(repo_root: Path, case_id: Optional[str] = None) -> Optional[TrialCredentials]:
    workbook_path = repo_root / "testmanager.xlsx"
    if not workbook_path.exists():
        logger.debug("Trial adapter: testmanager.xlsx not found at %s", workbook_path)
        return None
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        logger.warning("Trial adapter: openpyxl is required to read %s (%s)", workbook_path, exc)
        return None

    try:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Trial adapter: failed to open %s (%s)", workbook_path, exc)
        return None

    try:
        sheet = workbook["TestConfiguration"]
    except KeyError:
        logger.warning("Trial adapter: sheet 'TestConfiguration' not found in %s", workbook_path)
        workbook.close()
        return None

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {name: idx for idx, name in enumerate(headers)}

    def _cell(row, name):
        idx = header_index.get(name)
        if idx is None or idx >= len(row):
            return ""
        value = row[idx]
        if value is None:
            return ""
        return str(value).strip()

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    credentials: Optional[TrialCredentials] = None

    # Simplified rule: pick the FIRST row that has both a non-empty USERNAME and PASSWORD.
    # If none have both, fall back to first with any credential signal (USERNAME or PASSWORD or API_URL).
    first_partial: Optional[TrialCredentials] = None
    for row in rows:
        if not row:
            continue
        base_url = _cell(row, "API_URL")
        username = _cell(row, "USERNAME")
        password = _cell(row, "PASSWORD")
        if username and password:
            credentials = TrialCredentials(base_url=base_url, username=username, password=password)
            logger.info("Trial adapter: selected first full credential row (USERNAME+PASSWORD) from TestConfiguration")
            break
        if not first_partial and (username or password or base_url):
            first_partial = TrialCredentials(base_url=base_url, username=username, password=password)
    if not credentials and first_partial:
        credentials = first_partial
        logger.info("Trial adapter: fell back to first partial credential row from TestConfiguration")

    workbook.close()
    if not credentials:
        logger.warning("Trial adapter: no usable credentials found in TestConfiguration (%s)", workbook_path)
    return credentials


def trial_env_overrides(repo_root: Path, case_id: Optional[str] = None, spec_path: Optional[Path] = None) -> Dict[str, str]:
    """
    Build environment variable overrides for trial executions using credentials stored
    in the TestConfiguration sheet. This enables specs that rely on process.env to use
    consistent values without editing source files.
    """
    # Infer case id from spec content if not provided
    if not case_id and spec_path and spec_path.exists():
        try:
            source = spec_path.read_text(encoding="utf-8", errors="replace")
            titles = _extract_titles_from_source(source)
            # Pick the first ID-like title if present
            for t in titles:
                if _is_id_like(t):
                    case_id = t
                    break
        except Exception:  # noqa: BLE001
            pass

    credentials = load_trial_credentials(repo_root, case_id)
    if not credentials:
        return {}

    overrides: Dict[str, str] = {}
    # Always override a broad set of common env names so trial ignores .env creds
    if credentials.username:
        overrides["USERID"] = credentials.username
        overrides["USERNAME"] = credentials.username  # alias for tests using USERNAME
        overrides["TRIAL_USER"] = credentials.username
        overrides["TRIAL_USERNAME"] = credentials.username
        overrides["EMAIL"] = credentials.username  # some tests may treat username as email
    if credentials.password:
        overrides["PASSWORD"] = credentials.password
        overrides["TRIAL_PASSWORD"] = credentials.password
    if credentials.base_url:
        overrides["BASE_URL"] = credentials.base_url
        overrides["URL"] = credentials.base_url  # alias for tests using URL
        overrides["TRIAL_BASE_URL"] = credentials.base_url
        overrides["TRIAL_URL"] = credentials.base_url
    return overrides


def _replace_fill_call(source: str, pattern: str, replacement_value: str) -> Tuple[str, bool]:
    import re

    replaced = False

    def _replacer(match: "re.Match[str]") -> str:
        nonlocal replaced
        replaced = True
        prefix = match.group(1)
        suffix = match.group(2)
        return f"{prefix}{json.dumps(replacement_value)}{suffix}"

    updated = re.sub(pattern, _replacer, source, count=1)
    return updated, replaced


def _adjust_mfa_sequence(source: str) -> Tuple[str, bool]:
    import re

    changed = False

    replacements = [
        (
            r"(\s*)await\s+flow\.enterPasscode\.click\(\);\s*\n",
            "Trial adapter: user handles the MFA prompt manually during trial runs.",
        ),
        (
            r"(\s*)await\s+flow\.enterPasscode\.fill\([^;]*\);\s*\n",
            "Trial adapter: user enters the MFA passcode manually during trial runs.",
        ),
        (
            r"(\s*)await\s+flow\.verify\.click\(\);\s*\n",
            "Trial adapter: user submits verification manually during trial runs.",
        ),
    ]

    updated = source
    for pattern, message in replacements:
        def _replacer(match: "re.Match[str]") -> str:
            nonlocal changed
            changed = True
            indent = match.group(1)
            return f"{indent}// {message}\n"

        updated = re.sub(pattern, _replacer, updated, count=1)

    return updated, changed


def _inject_sign_in_pause(source: str) -> Tuple[str, bool]:
    import re

    if "page.waitForTimeout(60000)" in source:
        return source, False

    pattern = r"(await\s+flow\.signIn\.click\(\);\s*\n)"
    injected = False

    def _replacer(match: "re.Match[str]") -> str:
        nonlocal injected
        injected = True
        return (
            f"{match.group(1)}      // Trial adapter: wait to allow manual MFA passcode entry\n"
            "      await page.waitForTimeout(60000);\n"
        )

    updated = re.sub(pattern, _replacer, source, count=1)
    return updated, injected


def _ensure_navigation(source: str, base_url: str) -> Tuple[str, bool]:
    if "page.goto(" in source:
        return source, False
    marker = "    flow = new PageObject(page);"
    if marker not in source:
        return source, False
    navigation_line = f"    await page.goto({json.dumps(base_url)}, {{ waitUntil: 'load' }});"
    updated = source.replace(marker, f"{marker}\n{navigation_line}", 1)
    return updated, True


def adapt_spec_content_for_trial(source: str, repo_root: Path) -> Tuple[str, bool]:
    """Return transformed spec content for trial run; bool indicates change."""
    credentials = load_trial_credentials(repo_root)
    if not credentials:
        return source, False

    updated = source
    changed_any = False

    updated, user_changed = _replace_fill_call(
        updated,
        r"(await\s+flow\.userName\.fill\()\s*(?:['\"].*?['\"])(\);)",
        credentials.username,
    )
    updated, pass_changed = _replace_fill_call(
        updated,
        r"(await\s+flow\.password\.fill\()\s*(?:['\"].*?['\"])(\);)",
        credentials.password,
    )

    if not (user_changed or pass_changed):
        # No login steps detected; nothing to adapt.
        return source, False

    changed_any |= user_changed or pass_changed

    updated, nav_changed = _ensure_navigation(updated, credentials.base_url)
    changed_any |= nav_changed

    updated, pause_changed = _inject_sign_in_pause(updated)
    changed_any |= pause_changed

    updated, strip_changed = _adjust_mfa_sequence(updated)
    changed_any |= strip_changed

    return updated, changed_any


def prepare_trial_spec_path(spec_path: Path, repo_root: Path) -> Tuple[Path, Optional[Callable[[], None]]]:
    """Return a spec path to execute for trial runs; optionally provides cleanup callback."""
    try:
        original = spec_path.read_text(encoding="utf-8")
    except Exception as exc:  # noqa: BLE001
        logger.warning("Trial adapter: failed to read spec %s (%s)", spec_path, exc)
        return spec_path, None

    adapted, changed = adapt_spec_content_for_trial(original, repo_root)
    if not changed:
        return spec_path, None

    temp_dir = spec_path.parent
    try:
        fd, temp_path_str = tempfile.mkstemp(
            prefix=spec_path.stem + "_trial_",
            suffix=".spec.ts",
            dir=temp_dir,
        )
        os.close(fd)
        temp_path = Path(temp_path_str)
        temp_path.write_text(adapted, encoding="utf-8")
    except Exception as exc:  # noqa: BLE001
        logger.warning("Trial adapter: failed to write adapted spec for %s (%s)", spec_path, exc)
        return spec_path, None

    def _cleanup() -> None:
        try:
            temp_path.unlink(missing_ok=True)  # type: ignore[arg-type]
        except Exception:  # noqa: BLE001
            logger.debug("Trial adapter: failed to remove temp spec %s", temp_path)

    return temp_path, _cleanup
